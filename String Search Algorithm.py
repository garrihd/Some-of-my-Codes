#import needed library to open Excel files
from openpyxl import load_workbook


#initialize global variables
list_noun = []
list_noun_target = []
list_adverb = []
list_adverb_target = [] 

#create first workbook for SketchEngine Excel Corpus

wb = load_workbook("/home/garri/Downloads/concordance_preloaded_bulgarianNC_20210914104702.xlsx")
source = wb["concordance"]
for cell in source['C']:
    list_noun.append(cell.value)

#Create a list out of a set - to remove all duplicates       


list_noun = list(set(list_noun))
print("Inital length of noun list :", len(list_noun))
print(" ")


# Return all elements in list_a ending in "т"

for x in list_noun:
    x = str(x)
    if x[-1] == "т":
        list_noun_target.append(x)

for ch in range(len(list_noun_target)):
    list_noun_target[ch] = list_noun_target[ch].lower()

list_noun_target = list(set(list_noun_target))        
        
print(list_noun_target)
print(" ")
print("Length of target noun corpus is ", len(list_noun_target))
print("")
print("__________________________________")


#create second workbook for SketchEngine Excel Corpus

wb2 = load_workbook("/home/garri/Downloads/concordance_preloaded_bulgarianNC_20210918143804(adverbs).xlsx")
source = wb2["concordance"]
for cell1 in source['C']:
    
        list_adverb.append(cell1.value)
#Create a list out of a set - to remove all duplicates 
list_adverb = list(set(list_adverb))
print("")
print("Inital length of adverb list :", len(list_adverb))


# Return all elements in list_x with double "н"   
for ele in list_adverb:
    ele = str(ele)
    if "нн" in ele:
        
        list_adverb_target.append(ele)

        
#Turn all elements into lower case words

for i in range(len(list_adverb_target)):
    list_adverb_target[i] = list_adverb_target[i].lower()

#Create a list out of a set - to remove all duplicates

list_adverb_target = list(set(list_adverb_target))

print("")        
print(list_adverb_target)
print("")
print("Length of target adverb corpus is ", len(list_adverb_target))        